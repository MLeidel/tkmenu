'''
tkmenu.py
Python console program
Michael Leidel
builds tkinter menu from a xlsx file (template: menutpl.xlsx)
with these columns:
variable, label, command, accelerator, underline

NOTE: OUTPUTS CODE TO CONSOLE
'''

import sys
import os
import openpyxl

if (len(sys.argv)) > 1:
    infile = sys.argv[1]
else:
    print("missing input file name!")
    sys.exit()

if os.path.exists(infile):
    if not infile.endswith("xlsx"):
        print("must be an Excel (xlsx) file.")
        sys.exit()
else:
    print("cannot find file: " + infile)
    sys.exit()

# list (flds) index names for the (row)column values
nop, var, lbl, cmd, acc, und = 0, 1, 2, 3, 4, 5
MN_VAR = ""  # holds the menu object variable found in column 1 os ss
CALLBACKS = []  # holds the handler function name for later printout


def prt(strout):
    ''' Adjust leading tabs/spaces for output here '''
    print("        " + strout)


# Get the Excel workbook and spreadsheet.
# The Excel file is expected to be in the app's directory.

wb = openpyxl.load_workbook(infile)
sheet = wb.get_sheet_by_name('layout')  # Must be a Sheet titled 'layout' !
flds = []  # holds one row in ss

print("\n\n\n")
prt("menubar = Menu(root)")  # the menu code starts here

for rownum in range(2, sheet.max_row + 1):  # loop through each row after header

    flds.clear()  # clear list
    ROUT = ""  # holds output for each row of python/tkinter code
    flds.append("nop")  # zero element not used
    # load up the flds list with this row's columns values
    for c in range(1, 6):  # columns align with list index
        val = sheet.cell(row=rownum, column=c).value
        if val is None:
            flds.append("")
        else:
            flds.append(val)


    #  Process this row/columns values

    if flds[var] != MN_VAR:  # starts a new menu or submenu item
        if flds[var] == "submenu":
            ROUT = "submenu = Menu(" + MN_VAR +", tearoff=False)"
            MN_VAR = flds[var]
            prt(ROUT)
            # now continue to process the row
        elif MN_VAR == "submenu":  # finished with this submenu?
            ROUT = flds[var] + ".add_cascade(label=\""
            ROUT += flds[lbl][1:] + "\", menu=submenu"
            if flds[und] != "":
                ROUT += ", underline=" + flds[und]
            ROUT += ")"
            prt(ROUT)
            MN_VAR = flds[var]
            continue
        else:
            ROUT = flds[var] + " = Menu(menubar, tearoff=0)"
            MN_VAR = flds[var]
            prt(ROUT)
            # now continue to process the row

    if flds[lbl] == "separator":
        ROUT = flds[var] + ".add_separator()"
        prt(ROUT)
        continue

    if flds[lbl].startswith("_"):
        ROUT = "menubar.add_cascade(label=\""
        ROUT += flds[lbl][1:] + "\", menu=" + flds[var] + ")"
        prt(ROUT)
        continue

    # VAR + '.add_command(label="LBL", command=self.CMD, accelerator="ACC", underline=UND)'
    ROUT = flds[var] + ".add_command(label=\""
    ROUT += flds[lbl] + "\", command=self."
    ROUT += flds[cmd]
    CALLBACKS.append(flds[cmd])
    if flds[acc] != "":
        ROUT += ", accelerator=\"" + flds[acc] + "\""
    if flds[und] != "":
        ROUT += ", underline=" + flds[und]
    ROUT += ")"
    prt(ROUT)

prt("root.config(menu=menubar) # display the menu\n\n")  # here ends the menu code

# adjust tabsize (print spaces) below as needed
for item in CALLBACKS:
    print("    def %s(self):" % (item))
    print("        ''' docstring '''\n")
print("\n\n\n")
